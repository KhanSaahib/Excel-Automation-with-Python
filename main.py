"""
Project Title: Excel Automation with python
Author: Saddam Khan Ashna
This project performes data analysis on a simple excel data sheet using openpyxl python library.
"""

import openpyxl as open


inv_file = open.load_workbook("inventory.xlsx")
prod_table = inv_file["Sheet1"]

#Dictionaries to store the data
prod_per_suplier = {}
total_inv_of_suplier= {}
less_inv_prod = {}

for row in range(2, prod_table.max_row+1):
    suplier_name = prod_table.cell(row, 4).value
    inventory = prod_table.cell(row,2).value
    price = prod_table.cell(row,3).value
    product_id = prod_table.cell(row, 1).value
    inventory_price = prod_table.cell(row, 5)


    #Calculating number of product per supliers
    if suplier_name not in prod_per_suplier:
        prod_per_suplier[suplier_name] = 1
    else:
        value = prod_per_suplier[suplier_name]
        prod_per_suplier[suplier_name] = value +1

    #Calculating the total inventory value per suplier.
    if suplier_name not in total_inv_of_suplier:
        total_inv_of_suplier[suplier_name] = inventory * price
    else:
        var = total_inv_of_suplier.get(suplier_name)
        total_inv_of_suplier[suplier_name] += inventory*price

    #find the products with inventory less than 10
    if suplier_name not in less_inv_prod and inventory< 10:
        less_inv_prod[int(product_id)] = int(inventory)

    #adding values for the inventory price
    inventory_price.value = inventory * price



print(prod_per_suplier)
print(total_inv_of_suplier)
print(less_inv_prod)

#Saving the data into a new excel file
inv_file.save("invetory_updated.xlsx")